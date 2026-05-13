from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from .fixed_point import parse_fixed_format
from .model import Design


FIXED_HEADERS = ["signal", "fixed_format", "description"]


def _row_dict(headers: list[str], row: tuple[Any, ...]) -> dict[str, Any]:
    return {headers[i]: row[i] for i in range(min(len(headers), len(row)))}


def import_fixed_signals(path: str | Path, sheet_name: str = "fixed_signals") -> Design:
    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"missing sheet {sheet_name!r} in {path}")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return Design()
    headers = [str(v).strip() if v is not None else "" for v in rows[0]]
    required = {"signal", "fixed_format"}
    missing = required - set(headers)
    if missing:
        raise ValueError(f"fixed_signals missing required columns: {', '.join(sorted(missing))}")

    attrs = []
    for raw in rows[1:]:
        item = _row_dict(headers, raw)
        signal = str(item.get("signal") or "").strip()
        fixed_format = str(item.get("fixed_format") or "").strip()
        if not signal and not fixed_format:
            continue
        if not signal or not fixed_format:
            raise ValueError(f"incomplete fixed_signals row: {item}")
        description = item.get("description")
        attrs.append(parse_fixed_format(signal, fixed_format, str(description) if description else None))
    return Design(signal_attributes=attrs)


def create_fixed_signal_template(path: str | Path) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "fixed_signals"
    ws.append(FIXED_HEADERS)
    ws.append(["sample_in", "s(11,1)", "ADC sample"])
    ws.append(["filtered_out", "s(11,1)", "Filter output"])
    wb.save(path)

